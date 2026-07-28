#!/usr/bin/env python3
"""
Upload the acoustic.ge SEO audit workbook into a single Google Sheets tab
('Acoustic.ge - SEO & Content'), reproducing the local Excel layout
faithfully and legibly.

Why this reads the workbook with openpyxl instead of pandas:
the workbook relies on merged cells and title rows. pandas would treat the
first row as a header and turn every merged cell into NaN, which is what
produced the empty cells in the previous upload.

Georgian wording/grammar corrections live in seo_audit_text_fixes.py and are
applied both to the uploaded content and (optionally) back to the local file.
"""

import os
import sys
import shutil
import argparse
from datetime import datetime, timedelta, timezone

import gspread
import openpyxl

from seo_audit_text_fixes import fix_text

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CREDENTIALS_FILE = os.path.join(SCRIPT_DIR, 'Acoustic-Geovoice', 'credentials.json')
SPREADSHEET_ID = '1tDKgxcxPF8Jq151nMb6Wu_ziyOxkFATKSOquFKZrg94'
TAB_NAME = 'Acoustic.ge - SEO & Content'

# The single tab uses 8 columns so the widest table (the detailed report)
# fits natively; narrower sections merge their text across the spare columns.
TOTAL_COLS = 8
COLUMN_WIDTHS = [50, 135, 265, 335, 335, 335, 275, 125]

# Closing note appended after every section of the audit.
POSTSCRIPT_TITLE = 'P.S. — მოსალოდნელი ეფექტი'
POSTSCRIPT_TEXT = (
    'ტექნიკური SEO-ს და ინდექსაციის სრულად მოწესრიგებას შეუძლია საიტის ორგანული '
    'გაყიდვები მინიმუმ 30%-50%-ით (საბაზისო ტექნიკური გასწორებებით) და მაქსიმუმ '
    '150%-300%-ით ან მეტადაც კი გაზარდოს (თუ პროდუქტების სრული ბაზა ჩაერთვება '
    'Google-ის ძიებაში და ქართულენოვანი ოპტიმიზაცია დასრულდება).'
)

# Row kinds drive both layout and styling.
DOC_TITLE, DOC_META, SECTION, SUMMARY, HEADER, DATA, KV, PS = (
    'doc_title', 'doc_meta', 'section', 'summary', 'header', 'data', 'kv', 'ps')

SEVERITY_STYLES = {
    '🔴': ({'red': 0.988, 'green': 0.898, 'blue': 0.898}, {'red': 0.702, 'green': 0.098, 'blue': 0.098}),
    '🟠': ({'red': 1.000, 'green': 0.949, 'blue': 0.878}, {'red': 0.749, 'green': 0.420, 'blue': 0.047}),
    '🟡': ({'red': 1.000, 'green': 0.988, 'blue': 0.878}, {'red': 0.596, 'green': 0.478, 'blue': 0.000}),
    '🟢': ({'red': 0.898, 'green': 0.960, 'blue': 0.898}, {'red': 0.118, 'green': 0.502, 'blue': 0.200}),
    '✓': ({'red': 0.898, 'green': 0.960, 'blue': 0.898}, {'red': 0.118, 'green': 0.502, 'blue': 0.200}),
}

TEAL = {'red': 0.051, 'green': 0.400, 'blue': 0.400}
NAVY = {'red': 0.122, 'green': 0.306, 'blue': 0.471}
LIGHT_GRAY = {'red': 0.949, 'green': 0.949, 'blue': 0.949}
BAND_GRAY = {'red': 0.976, 'green': 0.980, 'blue': 0.984}
PS_YELLOW = {'red': 1.000, 'green': 0.976, 'blue': 0.878}
WHITE = {'red': 1.0, 'green': 1.0, 'blue': 1.0}
BORDER_GRAY = {'red': 0.80, 'green': 0.82, 'blue': 0.85}


def find_latest_seo_audit():
    """Return the most recently modified SEO audit workbook, or None."""
    candidates = []
    for name in os.listdir(SCRIPT_DIR):
        lowered = name.lower()
        if not lowered.endswith('.xlsx'):
            continue
        if 'seo' in lowered and 'audit' in lowered:
            path = os.path.join(SCRIPT_DIR, name)
            candidates.append((os.path.getmtime(path), path))
    if not candidates:
        return None
    return max(candidates)[1]


def read_workbook(path):
    """Read every sheet as raw rows of strings, preserving empty cells."""
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    for name in workbook.sheetnames:
        worksheet = workbook[name]
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            values = ['' if value is None else fix_text(str(value).strip())
                      for value in row]
            rows.append(values)
        # Drop fully empty trailing rows so blocks stay tight.
        while rows and not any(rows[-1]):
            rows.pop()
        sheets.append((fix_text(name), rows))
        print(f"   '{name}': {len(rows)} rows x "
              f"{max((len(r) for r in rows), default=0)} cols")
    return sheets


def cell(value, start, end, kind):
    """A single placement: value spanning columns [start, end] inclusive."""
    return {'value': value, 'start': start, 'end': end, 'kind': kind}


def at(row, index):
    return row[index] if index < len(row) else ''


def layout_overview(rows):
    """მიმოხილვა: titles span the full width, summary is label + count,
    the priorities table gets wide action and difficulty cells."""
    laid_out = []
    in_table = False
    for row in rows:
        first = at(row, 0)
        if not any(row):
            laid_out.append([])
            continue

        if first.startswith('SEO აუდიტი'):
            laid_out.append([cell(first, 0, 7, DOC_TITLE)])
        elif first.startswith('თარიღი:') or first.startswith('პროდუქტების რაოდენობა:'):
            laid_out.append([cell(first, 0, 7, DOC_META)])
        elif first in ('შეჯამება', 'გასაკეთებელი პრიორიტეტები'):
            in_table = first == 'გასაკეთებელი პრიორიტეტები'
            laid_out.append([cell(first, 0, 7, SECTION)])
        elif not in_table:
            # Summary line: label merged in A:E, count merged in F:H.
            count = next((at(row, i) for i in range(1, len(row)) if at(row, i)), '')
            laid_out.append([cell(first, 0, 4, SUMMARY),
                             cell(count, 5, 7, SUMMARY)])
        else:
            kind = HEADER if first == '#' else DATA
            laid_out.append([
                cell(first, 0, 0, kind),
                cell(at(row, 1), 1, 1, kind),
                cell(at(row, 2), 2, 5, kind),
                cell(at(row, 3), 6, 7, kind),
            ])
    return laid_out


def layout_detailed(rows):
    """დეტალური ანგარიში: already 8 columns, so keep it one-to-one."""
    laid_out = []
    for index, row in enumerate(rows):
        if not any(row):
            laid_out.append([])
            continue
        kind = HEADER if index == 0 else DATA
        laid_out.append([cell(at(row, col), col, col, kind)
                         for col in range(TOTAL_COLS)])
    return laid_out


def layout_good(rows):
    """რა კარგად არის: #, wide element text, status on the right."""
    laid_out = []
    for row in rows:
        first = at(row, 0)
        if not any(row):
            laid_out.append([])
            continue
        if first.startswith('🟢'):
            laid_out.append([cell(first, 0, 7, SECTION)])
            continue
        kind = HEADER if first == '#' else DATA
        laid_out.append([
            cell(first, 0, 0, kind),
            cell(at(row, 1), 1, 5, kind),
            cell(at(row, 2), 6, 7, kind),
        ])
    return laid_out


def layout_notes(rows):
    """შენიშვნა: label on the left, multi-line note across the rest."""
    laid_out = []
    for index, row in enumerate(rows):
        first = at(row, 0)
        if not any(row):
            laid_out.append([])
            continue
        if index == 0:
            laid_out.append([cell(first, 0, 7, SECTION)])
            continue
        laid_out.append([cell(first, 0, 1, KV),
                         cell(at(row, 1), 2, 7, KV)])
    return laid_out


LAYOUTS = {
    'მიმოხილვა': layout_overview,
    'დეტალური ანგარიში': layout_detailed,
    'რა კარგად არის': layout_good,
    'შენიშვნა': layout_notes,
}


def build_layout(sheets):
    """Flatten every sheet into one list of rows of placements."""
    combined = []
    for name, rows in sheets:
        builder = LAYOUTS.get(name, layout_detailed)
        if combined:
            combined.append([])  # breathing room between sections
        if name != 'მიმოხილვა':
            combined.append([cell(name, 0, 7, SECTION)])
        combined.extend(builder(rows))
    while combined and not combined[-1]:
        combined.pop()
    combined.append([])
    combined.append([cell(POSTSCRIPT_TITLE, 0, 7, SECTION)])
    combined.append([cell(POSTSCRIPT_TEXT, 0, 7, PS)])
    return combined


def to_values(layout):
    """Turn placements into the plain 2D grid gspread uploads."""
    grid = []
    for row in layout:
        line = [''] * TOTAL_COLS
        for placement in row:
            line[placement['start']] = placement['value']
        grid.append(line)
    return grid


def severity_style(text):
    """Return (background, foreground) for severity/status markers."""
    stripped = str(text).strip()
    for marker, style in SEVERITY_STYLES.items():
        if stripped.startswith(marker):
            return style
    return None


def text_format(bold=False, size=10, color=None):
    fmt = {'bold': bold, 'fontSize': size,
           'fontFamily': 'Noto Sans Georgian'}
    if color:
        fmt['foregroundColor'] = color
    return fmt


def build_requests(layout, sheet_id):
    """Return (style_requests, height_requests).

    Heights are returned separately because they must be applied *after* the
    row auto-fit pass, otherwise auto-fit would shrink the banner rows.
    """
    requests = []
    heights = []

    def set_height(row_index, pixels):
        heights.append({'updateDimensionProperties': {
            'range': {'sheetId': sheet_id, 'dimension': 'ROWS',
                      'startIndex': row_index, 'endIndex': row_index + 1},
            'properties': {'pixelSize': pixels}, 'fields': 'pixelSize',
        }})

    def rng(row_index, start, end):
        return {'sheetId': sheet_id,
                'startRowIndex': row_index, 'endRowIndex': row_index + 1,
                'startColumnIndex': start, 'endColumnIndex': end + 1}

    def repeat(row_index, start, end, cell_format, fields):
        requests.append({'repeatCell': {
            'range': rng(row_index, start, end),
            'cell': {'userEnteredFormat': cell_format},
            'fields': f'userEnteredFormat({fields})',
        }})

    borders = {side: {'style': 'SOLID', 'color': BORDER_GRAY}
               for side in ('top', 'bottom', 'left', 'right')}
    data_row_number = 0

    for row_index, row in enumerate(layout):
        if not row:
            continue

        for placement in row:
            start, end, kind = placement['start'], placement['end'], placement['kind']
            value = placement['value']

            if end > start:
                requests.append({'mergeCells': {
                    'range': rng(row_index, start, end),
                    'mergeType': 'MERGE_ALL',
                }})

            if kind == DOC_TITLE:
                repeat(row_index, start, end, {
                    'backgroundColor': TEAL,
                    'textFormat': text_format(True, 16, WHITE),
                    'horizontalAlignment': 'CENTER',
                    'verticalAlignment': 'MIDDLE',
                    'wrapStrategy': 'WRAP',
                }, 'backgroundColor,textFormat,horizontalAlignment,verticalAlignment,wrapStrategy')
                set_height(row_index, 48)

            elif kind == DOC_META:
                repeat(row_index, start, end, {
                    'backgroundColor': LIGHT_GRAY,
                    'textFormat': text_format(False, 10),
                    'horizontalAlignment': 'CENTER',
                    'verticalAlignment': 'MIDDLE',
                    'wrapStrategy': 'WRAP',
                }, 'backgroundColor,textFormat,horizontalAlignment,verticalAlignment,wrapStrategy')

            elif kind == SECTION:
                repeat(row_index, start, end, {
                    'backgroundColor': NAVY,
                    'textFormat': text_format(True, 12, WHITE),
                    'horizontalAlignment': 'LEFT',
                    'verticalAlignment': 'MIDDLE',
                    'wrapStrategy': 'WRAP',
                    'padding': {'left': 10},
                }, 'backgroundColor,textFormat,horizontalAlignment,verticalAlignment,wrapStrategy,padding')
                set_height(row_index, 36)

            elif kind == SUMMARY:
                style = severity_style(value)
                background = style[0] if style else LIGHT_GRAY
                foreground = style[1] if style else None
                is_count = start == 5
                repeat(row_index, start, end, {
                    'backgroundColor': background,
                    'textFormat': text_format(True, 11, foreground),
                    'horizontalAlignment': 'CENTER' if is_count else 'LEFT',
                    'verticalAlignment': 'MIDDLE',
                    'wrapStrategy': 'WRAP',
                    'borders': borders,
                    'padding': {'left': 10, 'right': 10},
                }, 'backgroundColor,textFormat,horizontalAlignment,verticalAlignment,wrapStrategy,borders,padding')

            elif kind == HEADER:
                repeat(row_index, start, end, {
                    'backgroundColor': NAVY,
                    'textFormat': text_format(True, 10, WHITE),
                    'horizontalAlignment': 'CENTER',
                    'verticalAlignment': 'MIDDLE',
                    'wrapStrategy': 'WRAP',
                    'borders': borders,
                }, 'backgroundColor,textFormat,horizontalAlignment,verticalAlignment,wrapStrategy,borders')

            elif kind == PS:
                repeat(row_index, start, end, {
                    'backgroundColor': PS_YELLOW,
                    'textFormat': text_format(False, 11),
                    'horizontalAlignment': 'LEFT',
                    'verticalAlignment': 'MIDDLE',
                    'wrapStrategy': 'WRAP',
                    'borders': borders,
                    'padding': {'left': 12, 'right': 12, 'top': 10, 'bottom': 10},
                }, 'backgroundColor,textFormat,horizontalAlignment,verticalAlignment,wrapStrategy,borders,padding')

            elif kind == KV:
                is_label = start == 0
                repeat(row_index, start, end, {
                    'backgroundColor': LIGHT_GRAY if is_label else WHITE,
                    'textFormat': text_format(is_label, 10),
                    'horizontalAlignment': 'LEFT',
                    'verticalAlignment': 'TOP',
                    'wrapStrategy': 'WRAP',
                    'borders': borders,
                    'padding': {'left': 10, 'right': 10, 'top': 6, 'bottom': 6},
                }, 'backgroundColor,textFormat,horizontalAlignment,verticalAlignment,wrapStrategy,borders,padding')

            else:  # DATA
                style = severity_style(value)
                if style:
                    background, foreground, bold = style[0], style[1], True
                else:
                    background = BAND_GRAY if data_row_number % 2 else WHITE
                    foreground, bold = None, False
                numeric = value.isdigit()
                repeat(row_index, start, end, {
                    'backgroundColor': background,
                    'textFormat': text_format(bold, 10, foreground),
                    'horizontalAlignment': 'CENTER' if (numeric or style) else 'LEFT',
                    'verticalAlignment': 'TOP',
                    'wrapStrategy': 'WRAP',
                    'borders': borders,
                    'padding': {'left': 8, 'right': 8, 'top': 6, 'bottom': 6},
                }, 'backgroundColor,textFormat,horizontalAlignment,verticalAlignment,wrapStrategy,borders,padding')

        if row and row[0]['kind'] == DATA:
            data_row_number += 1
        if not row:
            set_height(row_index, 12)

    for index, width in enumerate(COLUMN_WIDTHS):
        requests.append({'updateDimensionProperties': {
            'range': {'sheetId': sheet_id, 'dimension': 'COLUMNS',
                      'startIndex': index, 'endIndex': index + 1},
            'properties': {'pixelSize': width}, 'fields': 'pixelSize',
        }})

    requests.append({'updateSheetProperties': {
        'properties': {'sheetId': sheet_id,
                       'gridProperties': {'frozenRowCount': 3}},
        'fields': 'gridProperties.frozenRowCount',
    }})

    return requests, heights


def apply_local_fixes(path):
    """Write the Georgian corrections back into the local workbook."""
    workbook = openpyxl.load_workbook(path)
    changes = 0
    for name in workbook.sheetnames:
        worksheet = workbook[name]
        for row in worksheet.iter_rows():
            for target in row:
                if isinstance(target.value, str):
                    fixed = fix_text(target.value)
                    if fixed != target.value:
                        target.value = fixed
                        changes += 1
    if changes:
        backup = path + '.bak'
        if not os.path.exists(backup):
            shutil.copy2(path, backup)
            print(f"   Backup created: {backup}")
        workbook.save(path)
    print(f"   Corrected {changes} cell(s) in the local workbook")
    return changes


def chunked(requests, size=100):
    for start in range(0, len(requests), size):
        yield requests[start:start + size]


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--file', help='Path to the SEO audit .xlsx')
    parser.add_argument('--skip-local-fixes', action='store_true',
                        help='Do not rewrite the local workbook')
    args = parser.parse_args()

    print('=== SEO AUDIT -> GOOGLE SHEETS ===\n')

    excel_file = args.file or find_latest_seo_audit()
    if not excel_file or not os.path.exists(excel_file):
        print('ERROR: SEO audit workbook not found.')
        return False
    print(f'1. Workbook: {excel_file}')

    if not args.skip_local_fixes:
        print('\n2. Applying Georgian corrections to the local workbook...')
        apply_local_fixes(excel_file)
    else:
        print('\n2. Local corrections skipped (--skip-local-fixes)')

    print('\n3. Reading sheets...')
    sheets = read_workbook(excel_file)
    if not sheets:
        print('ERROR: workbook has no readable sheets.')
        return False

    print('\n4. Building single-tab layout...')
    layout = build_layout(sheets)
    values = to_values(layout)
    print(f'   {len(values)} rows x {TOTAL_COLS} columns')

    print('\n5. Authenticating...')
    if not os.path.exists(CREDENTIALS_FILE):
        print(f'   ERROR: credentials not found at {CREDENTIALS_FILE}')
        return False
    try:
        client = gspread.service_account(filename=CREDENTIALS_FILE)
    except Exception as error:
        print(f'   ERROR: authentication failed: {error}')
        return False
    print('   OK')

    print(f"\n6. Opening tab '{TAB_NAME}'...")
    try:
        spreadsheet = client.open_by_key(SPREADSHEET_ID)
    except Exception as error:
        print(f'   ERROR: cannot open spreadsheet: {error}')
        return False
    try:
        worksheet = spreadsheet.worksheet(TAB_NAME)
        print('   Found existing tab')
    except gspread.WorksheetNotFound:
        worksheet = spreadsheet.add_worksheet(
            title=TAB_NAME, rows=len(values) + 20, cols=TOTAL_COLS)
        print('   Created tab')

    sheet_id = worksheet._properties['sheetId']

    print('\n7. Resetting tab (values, merges, formatting)...')
    reset = [
        {'updateSheetProperties': {
            'properties': {'sheetId': sheet_id, 'gridProperties': {
                'rowCount': max(len(values) + 10, 20),
                'columnCount': TOTAL_COLS,
                'frozenRowCount': 0}},
            'fields': 'gridProperties(rowCount,columnCount,frozenRowCount)'}},
        {'unmergeCells': {'range': {'sheetId': sheet_id}}},
        {'repeatCell': {
            'range': {'sheetId': sheet_id},
            'cell': {'userEnteredFormat': {}},
            'fields': 'userEnteredFormat'}},
    ]
    try:
        spreadsheet.batch_update({'requests': reset})
        worksheet.clear()
        print('   OK')
    except Exception as error:
        print(f'   ERROR: reset failed: {error}')
        return False

    print(f'\n8. Uploading {len(values)} rows...')
    try:
        worksheet.update(values=values, range_name='A1')
        print('   OK')
    except Exception as error:
        print(f'   ERROR: upload failed: {error}')
        return False

    print('\n9. Applying layout and styling...')
    requests, heights = build_requests(layout, sheet_id)
    try:
        for batch in chunked(requests):
            spreadsheet.batch_update({'requests': batch})
        # Auto-fit wrapped text first, then pin the banner rows.
        spreadsheet.batch_update({'requests': [{'autoResizeDimensions': {
            'dimensions': {'sheetId': sheet_id, 'dimension': 'ROWS',
                           'startIndex': 0, 'endIndex': len(values)}}}]})
        for batch in chunked(heights):
            spreadsheet.batch_update({'requests': batch})
        print(f'   Applied {len(requests) + len(heights)} formatting request(s)')
    except Exception as error:
        print(f'   WARNING: styling incomplete: {error}')

    stamp = datetime.now(timezone(timedelta(hours=4))).strftime('%Y-%m-%d %H:%M:%S')
    print('\n=== DONE ===')
    print(f'   Sections: {", ".join(name for name, _ in sheets)}')
    print(f'   Rows: {len(values)}   Tab: {TAB_NAME}')
    print(f'   Finished: {stamp} (Tbilisi)')
    return True


if __name__ == '__main__':
    sys.exit(0 if main() else 1)
